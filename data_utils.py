from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from io import BytesIO
from telegram import InputFile
from constants import *
import json


def load_data():
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {"specialties": {}}
    except json.JSONDecodeError:
        return {"specialties": {}}


def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


RESULTS_FILE = "test_results.json"


def load_results():
    if not os.path.exists(RESULTS_FILE):
        return []
    with open(RESULTS_FILE, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
            if isinstance(data, list):
                return data
            else:
                return []
        except json.JSONDecodeError:
            return []


def save_results(data):
    with open(RESULTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def add_result(fio, specialty, correct, total, username="—", user_id="—"):
    result_file = "test_results.json"
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    new_result = {
        "fio": fio,
        "username": username,
        "user_id": user_id,
        "specialty": specialty,
        "correct": correct,
        "total": total,
        "timestamp": timestamp
    }

    if os.path.exists(result_file):
        with open(result_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = []

    data.append(new_result)

    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


async def send_full_report(update, context):
    results = load_results()
    if not results:
        await update.message.reply_text("📭 Нет доступных результатов.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аттестации"
    headers = ["№", "ФИО", "Специальность", "Username", "Telegram ID", "Дата", "Правильных", "Всего"]
    ws.append(headers)

    for i, r in enumerate(results, start=1):
        ws.append([
            i,
            r.get("fio", "—"),
            r.get("specialty", "—"),
            r.get("username", "—"),
            r.get("user_id", "—"),
            r.get("timestamp", "—"),
            r.get("correct", "—"),
            r.get("total", "—")
        ])

    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    num_cols = [1, 5, 7, 8]
    for col_idx in num_cols:
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            cell.alignment = center_align

    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        if col_idx == 3:
            ws.column_dimensions[col_letter].width = max(max_length + 5, 25)
        else:
            ws.column_dimensions[col_letter].width = max_length + 5

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    await update.message.reply_document(InputFile(output, filename="аттестации.xlsx"))
