from math import floor

from telegram import ReplyKeyboardMarkup
from constants import *
from data_utils import load_results
from data_utils import load_data


async def send_full_report(update, context, role="admin"):
    results = load_results()
    if not results:
        await update.message.reply_text("📭 Нет доступных результатов.")
        return ADMIN_MENU if role == "admin" else MENTOR_MENU

    data = load_data()
    specialties = sorted(data.get("specialties", {}).keys())
    context.user_data["report_specialties"] = specialties
    context.user_data["report_role"] = role

    specialties_text = f"📋 Доступные специальности ({role}):\n\n" + \
        "\n".join([f"{i + 1}. {spec}" for i, spec in enumerate(specialties)]) + \
        "\n\nВведите номер специальности или нажмите «🔙 Назад»:"

    await update.message.reply_text(
        specialties_text,
        reply_markup=ReplyKeyboardMarkup([["🔙 Назад"]], resize_keyboard=True, one_time_keyboard=True)
    )
    return SELECT_SPECIALTY_FOR_REPORT


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from telegram import InputFile


async def handle_selected_specialty_report(update, context):
    text = update.message.text.strip()
    role = context.user_data.get("report_role", "admin")

    if text == "🔙 Назад":
        if role == "admin":
            from handlers.admin import admin_menu
            return await admin_menu(update, context)
        else:
            from handlers.mentor import mentor_menu
            return await mentor_menu(update, context)

    specialties = context.user_data.get("report_specialties", [])
    from data_utils import load_results
    results = load_results()

    try:
        index = int(text) - 1
        if index < 0 or index >= len(specialties):
            raise ValueError
        specialty = specialties[index]
    except:
        await update.message.reply_text("⚠️ Некорректный номер. Попробуйте ещё раз.")
        return SELECT_SPECIALTY_FOR_REPORT

    print(">> selected specialty:", specialty)
    print(">> all specialties in results:")
    for r in results:
        print("   -", r["specialty"])
    filtered = [
        r for r in results
        if r["specialty"] == specialty or r["specialty"].startswith(specialty + "::")
    ]

    if not filtered:
        filtered = [r for r in results if r["specialty"].startswith(specialty + "::")]

        if not filtered:
            await update.message.reply_text(
                f"📭 По специальности «{specialty}» и её подвидам никто ещё не проходил аттестацию.\n\n"
                f"Введите другой номер или нажмите «🔙 Назад»:"
            )
            return SELECT_SPECIALTY_FOR_REPORT
        else:
            await update.message.reply_text(
                f"📭 По специальности «{specialty}» никто ещё не проходил аттестацию.\n\n"
                f"Введите другой номер или нажмите «🔙 Назад»:"
            )
            return SELECT_SPECIALTY_FOR_REPORT

    wb = Workbook()
    ws = wb.active
    ws.title = "Аттестации"

    headers = [
        "№", "ФИО", "Профессия", "Структурное подразделение",
        "Обучен по специальности", "Username", "Telegram ID",
        "Дата аттестации", "Правильных", "Всего", "Итоговый результат"
    ]
    ws.append(headers)

    filtered.sort(key=lambda r: r.get("fio", "").lower())

    for i, r in enumerate(filtered, start=1):
        fio = r.get("fio", "—")
        specialty = r.get("specialty", "—")
        username = r.get("username", "—")
        user_id = r.get("user_id", "—")
        timestamp = r.get("timestamp", "—")
        correct = r.get("correct", 0)
        total = r.get("total", 0)

        errors = total - correct

        if total == 0 or errors > 3:
            result = "Пересдача"
        elif errors == 3:
            result = "7"
        else:
            result = str(round(correct * 10 / total))

        ws.append([
            i,
            fio,
            "",
            "",
            specialty,
            username,
            user_id,
            timestamp,
            correct,
            total,
            result
        ])

    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_length + 5, 20)

        for cell in col:
            cell.alignment = center_align

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"Аттестация_{specialty.replace('::', '_')}.xlsx"

    await update.message.reply_document(
        InputFile(output, filename=file_name),
        caption=f"📊 Сводный отчёт по: {specialty}"
    )

    await update.message.reply_text("📨 Можете ввести следующий номер специальности или нажать «🔙 Назад»:")
    return SELECT_SPECIALTY_FOR_REPORT

